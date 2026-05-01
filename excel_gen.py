"""
生成各省电价 Excel 报告，包含四个工作表：
  Sheet1 汇总   - 所有省份燃煤基准价 + 最新现货周均价
  Sheet2 现货   - 有现货市场省份的日前/实时均价（最近一期周数据）
  Sheet3 机制电价 - 各省新能源机制电价（风电/光伏）
  Sheet4 分时电价 - 各省峰平谷电价
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from provinces import PROVINCES, MECHANISM_PRICES, TIME_OF_USE_PRICES, MONTHLY_CONTRACT_PRICES

# ── 配色 ────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # 深蓝标题行
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "2E75B6"   # 次级标题
C_GRID_ODD    = "EBF3FB"   # 斑马纹奇行
C_GRID_EVEN   = "FFFFFF"
C_SPOT_HIGH   = "FF0000"   # 现货高价警示（>500 元/MWh）
C_SPOT_MED    = "FFC000"   # 中等价格
C_SPOT_LOW    = "70AD47"   # 低价
C_ACCENT      = "D6E4F0"   # 小节标题背景


def _hdr(text, bold=True, size=11, color=C_HEADER_FG, bg=C_HEADER_BG):
    return {
        "value": text,
        "font": Font(name="微软雅黑", bold=bold, size=size, color=color),
        "fill": PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _cell(value, bold=False, size=10, color="000000", bg=None, align="center", fmt=None, wrap=False):
    d = {
        "value": value,
        "font": Font(name="微软雅黑", bold=bold, size=size, color=color),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap),
        "border": _thin_border(),
    }
    if bg:
        d["fill"] = PatternFill("solid", fgColor=bg)
    if fmt:
        d["number_format"] = fmt
    return d


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply(ws, row, col, props: dict):
    c = ws.cell(row=row, column=col)
    c.value = props.get("value")
    if "font" in props:
        c.font = props["font"]
    if "fill" in props:
        c.fill = props["fill"]
    if "alignment" in props:
        c.alignment = props["alignment"]
    if "border" in props:
        c.border = props["border"]
    if "number_format" in props:
        c.number_format = props["number_format"]
    return c


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet 1：汇总 ────────────────────────────────────────
def build_summary(ws, spot_data: dict):
    ws.title = "汇总"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 36

    update_time = datetime.now().strftime("%Y年%m月%d日 %H:%M")
    week_label = spot_data.get("week", "—")

    # 大标题
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=f"全国各省电价汇总  |  数据更新：{update_time}")
    c.font = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # 列标题
    headers = ["电网分区", "省份", "燃煤基准价\n(元/kWh)",
               f"现货日前均价\n(元/MWh)\n{week_label}",
               "现货市场", "风电机制电价\n(元/kWh)", "光伏机制电价\n(元/kWh)", "备注"]
    widths   = [10, 10, 14, 18, 10, 16, 16, 18]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 42

    # 数据行
    spot_map = {d["province"]: d for d in spot_data.get("data", [])}
    prev_grid = None
    row = 3
    for p in PROVINCES:
        bg = C_GRID_ODD if row % 2 == 1 else C_GRID_EVEN

        # 电网分区（合并同组首行）
        if p["grid"] != prev_grid:
            prev_grid = p["grid"]
        _apply(ws, row, 1, _cell(p["grid"], bg=bg, bold=True, color="1F4E79"))

        _apply(ws, row, 2, _cell(p["name"], bg=bg, align="left"))
        _apply(ws, row, 3, _cell(p["coal_benchmark"], bg=bg, fmt='0.0000'))
        ws.cell(row=row, column=3).number_format = "0.0000"

        # 现货均价
        sd = spot_map.get(p["name"])
        if sd:
            avg = sd.get("da_avg")
            val = f"{avg:.1f}" if avg is not None else "—"
            # 颜色警示
            price_bg = bg
            if avg and avg > 500:
                price_bg = "FFD7D7"
            elif avg and avg > 300:
                price_bg = "FFF2CC"
            _apply(ws, row, 4, _cell(val, bg=price_bg, color="C00000" if (avg and avg > 500) else "000000"))
        else:
            _apply(ws, row, 4, _cell("—", bg=bg, color="999999"))

        spot_label = "✔ 已开通" if p["spot"] else "—"
        spot_color = "00B050" if p["spot"] else "999999"
        _apply(ws, row, 5, _cell(spot_label, bg=bg, color=spot_color))

        mech = MECHANISM_PRICES.get(p["name"], {})
        _apply(ws, row, 6, _cell(mech.get("wind", "—"), bg=bg, fmt='0.0000'))
        _apply(ws, row, 7, _cell(mech.get("solar", "—"), bg=bg, fmt='0.0000'))
        _apply(ws, row, 8, _cell(mech.get("note", ""), bg=bg, align="left", wrap=True))

        ws.row_dimensions[row].height = 18
        row += 1

    # 说明行
    ws.merge_cells(f"A{row}:H{row}")
    note = ws.cell(row=row, column=1,
                   value="说明：燃煤基准价来源国家发改委；机制电价为\"136号文\"新能源年度竞价出清结果；"
                         "现货均价为日前市场算术均价，蒙西为发电侧价格；标注\"约值\"或\"参考值\"表示非官方公布精确数据。")
    note.font = Font(name="微软雅黑", size=9, color="595959", italic=True)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


# ── Sheet 2：现货市场 ────────────────────────────────────
def build_spot(ws, spot_data: dict):
    ws.title = "现货市场"
    ws.sheet_view.showGridLines = False

    week_label = spot_data.get("week", "—")
    source_url = spot_data.get("source_url", "")

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"电力现货市场价格  |  统计周期：{week_label}")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["省份", "日前均价(元/MWh)", "日前最低(元/MWh)", "日前最高(元/MWh)", "实时均价(元/MWh)", "折算(元/kWh)"]
    widths   = [10, 18, 18, 18, 18, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 36

    for i, d in enumerate(spot_data.get("data", []), start=3):
        bg = C_GRID_ODD if i % 2 == 1 else C_GRID_EVEN
        avg = d.get("da_avg")
        if avg and avg > 500:
            bg = "FFE0E0"
        elif avg and avg > 350:
            bg = "FFF7E0"

        _apply(ws, i, 1, _cell(d["province"], bg=bg, bold=True))
        _apply(ws, i, 2, _cell(d.get("da_avg"), bg=bg, fmt="0.00"))
        ws.cell(row=i, column=2).number_format = "0.00"
        _apply(ws, i, 3, _cell(d.get("da_min"), bg=bg, fmt="0.00"))
        ws.cell(row=i, column=3).number_format = "0.00"
        _apply(ws, i, 4, _cell(d.get("da_max"), bg=bg, fmt="0.00"))
        ws.cell(row=i, column=4).number_format = "0.00"
        rt = d.get("rt_avg")
        _apply(ws, i, 5, _cell(rt, bg=bg, fmt="0.00"))
        if rt:
            ws.cell(row=i, column=5).number_format = "0.00"
        # 折算 元/kWh = 元/MWh ÷ 1000
        kwh_val = round(avg / 1000, 4) if avg else None
        _apply(ws, i, 6, _cell(kwh_val, bg=bg, fmt="0.0000"))
        if kwh_val:
            ws.cell(row=i, column=6).number_format = "0.0000"
        ws.row_dimensions[i].height = 18

    # 色阶说明
    note_row = 3 + len(spot_data.get("data", []))
    ws.merge_cells(f"A{note_row}:F{note_row}")
    n = ws.cell(row=note_row, column=1,
                value=f"数据来源：新浪财经电力现货市场价格周报  {source_url}")
    n.font = Font(name="微软雅黑", size=9, color="595959", italic=True)
    n.alignment = Alignment(horizontal="left", wrap_text=True)


# ── Sheet 3：机制电价 ────────────────────────────────────
def build_mechanism(ws):
    ws.title = "机制电价（新能源）"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="各省新能源机制电价  |  依据：国家发改委\"136号文\"  |  数据年份：2025-2026")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["电网分区", "省份", "风电机制电价(元/kWh)", "光伏机制电价(元/kWh)", "价差(风-光)", "备注"]
    widths   = [10, 10, 20, 20, 14, 28]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 36

    # 按电网分区排序
    grid_order = ["华北", "东北", "华东", "华中", "南方", "山东", "西北", "西藏"]
    province_grid = {p["name"]: p["grid"] for p in PROVINCES}

    row = 3
    for grid in grid_order:
        provs = [p for p in PROVINCES if p["grid"] == grid]
        for p in provs:
            mech = MECHANISM_PRICES.get(p["name"])
            if not mech:
                continue
            bg = C_GRID_ODD if row % 2 == 1 else C_GRID_EVEN
            wind = mech.get("wind")
            solar = mech.get("solar")
            diff = round(wind - solar, 4) if (wind and solar) else None

            _apply(ws, row, 1, _cell(grid, bg=bg, bold=True, color="1F4E79"))
            _apply(ws, row, 2, _cell(p["name"], bg=bg, align="left"))
            _apply(ws, row, 3, _cell(wind, bg=bg, fmt="0.0000"))
            ws.cell(row=row, column=3).number_format = "0.0000"
            _apply(ws, row, 4, _cell(solar, bg=bg, fmt="0.0000"))
            ws.cell(row=row, column=4).number_format = "0.0000"
            diff_bg = "E2EFDA" if diff and diff > 0 else ("FFE0E0" if diff and diff < 0 else bg)
            _apply(ws, row, 5, _cell(diff, bg=diff_bg, fmt="0.0000"))
            if diff:
                ws.cell(row=row, column=5).number_format = "0.0000"
            _apply(ws, row, 6, _cell(mech.get("note", ""), bg=bg, align="left", wrap=True))
            ws.row_dimensions[row].height = 18
            row += 1


# ── Sheet 4：分时电价 ────────────────────────────────────
def build_tou(ws):
    ws.title = "分时电价"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="各省分时电价（一般工商业 1-10kV）  |  数据年份：2026年1月")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["电网分区", "省份", "高峰电价(元/kWh)", "平段电价(元/kWh)", "低谷电价(元/kWh)", "峰谷价差", "备注"]
    widths   = [10, 10, 18, 18, 18, 12, 18]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 36

    grid_order = ["华北", "东北", "华东", "华中", "南方", "山东", "西北", "西藏"]
    row = 3
    for grid in grid_order:
        provs = [p for p in PROVINCES if p["grid"] == grid]
        for p in provs:
            tou = TIME_OF_USE_PRICES.get(p["name"])
            if not tou:
                continue
            bg = C_GRID_ODD if row % 2 == 1 else C_GRID_EVEN
            peak   = tou.get("peak")
            flat   = tou.get("flat")
            valley = tou.get("valley")
            diff = round(peak - valley, 4) if (peak and valley) else None

            _apply(ws, row, 1, _cell(grid, bg=bg, bold=True, color="1F4E79"))
            _apply(ws, row, 2, _cell(p["name"], bg=bg, align="left"))
            peak_bg = "FFD7D7" if (peak and peak > 1.2) else bg
            _apply(ws, row, 3, _cell(peak, bg=peak_bg, fmt="0.0000",
                                     color="C00000" if (peak and peak > 1.2) else "000000"))
            ws.cell(row=row, column=3).number_format = "0.0000"
            _apply(ws, row, 4, _cell(flat, bg=bg, fmt="0.0000"))
            ws.cell(row=row, column=4).number_format = "0.0000"
            valley_bg = "E2EFDA" if (valley and valley < 0.15) else bg
            _apply(ws, row, 5, _cell(valley, bg=valley_bg, fmt="0.0000",
                                     color="375623" if (valley and valley < 0.15) else "000000"))
            ws.cell(row=row, column=5).number_format = "0.0000"
            diff_bg = "FFF2CC" if (diff and diff > 1.0) else bg
            _apply(ws, row, 6, _cell(diff, bg=diff_bg, fmt="0.0000"))
            if diff:
                ws.cell(row=row, column=6).number_format = "0.0000"
            _apply(ws, row, 7, _cell(tou.get("note", ""), bg=bg, align="left"))
            ws.row_dimensions[row].height = 18
            row += 1

    note_row = row
    ws.merge_cells(f"A{note_row}:G{note_row}")
    n = ws.cell(row=note_row, column=1, value="红色：高峰价格>1.2元/kWh；绿色：低谷价格<0.15元/kWh；黄色：峰谷价差>1.0元/kWh")
    n.font = Font(name="微软雅黑", size=9, color="595959", italic=True)


# ── Sheet 5：中长期月度电价 ──────────────────────────────────
def build_monthly_contract(ws):
    ws.title = "中长期月度电价"
    ws.sheet_view.showGridLines = False

    # 取数据期标签（从第一条数据获取）
    periods = {v["period"] for v in MONTHLY_CONTRACT_PRICES.values()}
    period_label = sorted(periods)[-1] if periods else "—"

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1,
                value=f"各省中长期月度成交均价  |  数据期：{period_label}  |  标注\"估算\"表示非官方数据")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["电网分区", "省份", "月度中长期均价\n(元/kWh)", "燃煤基准价\n(元/kWh)", "价差\n(元/kWh)", "数据来源", "备注"]
    widths   = [10, 10, 18, 16, 14, 20, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 42

    grid_order = ["华北", "东北", "华东", "华中", "南方", "山东", "西北", "西藏"]
    row = 3
    for grid in grid_order:
        for p in [x for x in PROVINCES if x["grid"] == grid]:
            bg = C_GRID_ODD if row % 2 == 1 else C_GRID_EVEN
            mc = MONTHLY_CONTRACT_PRICES.get(p["name"])
            benchmark = p["coal_benchmark"]

            _apply(ws, row, 1, _cell(grid, bg=bg, bold=True, color="1F4E79"))
            _apply(ws, row, 2, _cell(p["name"], bg=bg, align="left"))

            if mc:
                price = mc["price"]
                diff  = round(price - benchmark, 4)
                diff_bg = "E2EFDA" if diff > 0 else ("FFE0E0" if diff < -0.02 else bg)
                diff_color = "375623" if diff > 0 else ("C00000" if diff < -0.02 else "000000")
                _apply(ws, row, 3, _cell(price, bg=bg, fmt="0.0000"))
                ws.cell(row=row, column=3).number_format = "0.0000"
                _apply(ws, row, 4, _cell(benchmark, bg=bg, fmt="0.0000"))
                ws.cell(row=row, column=4).number_format = "0.0000"
                _apply(ws, row, 5, _cell(diff, bg=diff_bg, color=diff_color, fmt="0.0000"))
                ws.cell(row=row, column=5).number_format = "0.0000"
                _apply(ws, row, 6, _cell(mc.get("source", ""), bg=bg, align="left"))
                _apply(ws, row, 7, _cell(mc.get("note", ""), bg=bg, align="left",
                                         color="C00000" if mc.get("note") == "估算" else "000000"))
            else:
                _apply(ws, row, 3, _cell("—", bg=bg, color="999999"))
                _apply(ws, row, 4, _cell(benchmark, bg=bg, fmt="0.0000"))
                ws.cell(row=row, column=4).number_format = "0.0000"
                for ci in [5, 6, 7]:
                    _apply(ws, row, ci, _cell("—", bg=bg, color="999999"))

            ws.row_dimensions[row].height = 18
            row += 1

    note_row = row
    ws.merge_cells(f"A{note_row}:G{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="绿色：成交均价高于燃煤基准价（发电侧有溢价）；红色：低于燃煤基准价2分以上（发电侧亏损压力）。"
                      "数据由 Remote Agent 每月初自动更新。")
    n.font = Font(name="微软雅黑", size=9, color="595959", italic=True)
    n.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 24


# ── Sheet 6：动力煤价 ────────────────────────────────────────
def build_coal_prices(ws, coal_data: dict):
    ws.title = "动力煤价"
    ws.sheet_view.showGridLines = False

    fetched_at = coal_data.get("fetched_at", "")[:10]
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value=f"主要产区/港口动力煤价格  |  数据日期：{fetched_at}  |  折算基准：发电煤耗305g/kWh标煤")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["地区", "煤种规格", "价格\n(元/吨)", "类型", "折算度电\n燃料成本\n(元/kWh)", "日期", "数据来源", "备注"]
    widths   = [16, 14, 12, 10, 16, 12, 16, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h, size=10))
        _set_col_width(ws, ci, w)
    ws.row_dimensions[2].height = 48

    for i, d in enumerate(coal_data.get("data", []), start=3):
        bg = C_GRID_ODD if i % 2 == 1 else C_GRID_EVEN
        price = d.get("price")

        # 折算度电燃料成本：price(元/吨) * 实际煤耗(g/kWh) / 1,000,000
        # 实际煤耗 = 305g/kWh标煤 * (7000 / 实际热值)
        spec = d.get("spec", "")
        kcal = 5500
        for k in [5500, 5000, 4500, 4000]:
            if str(k) in spec:
                kcal = k
                break
        actual_consumption = 305 * 7000 / kcal  # g/kWh
        fuel_cost = round(price * actual_consumption / 1_000_000, 4) if price else None

        _apply(ws, i, 1, _cell(d.get("region", ""), bg=bg, align="left", bold=True))
        _apply(ws, i, 2, _cell(spec, bg=bg, align="left"))
        _apply(ws, i, 3, _cell(price, bg=bg))
        _apply(ws, i, 4, _cell(d.get("type", ""), bg=bg))
        _apply(ws, i, 5, _cell(fuel_cost, bg=bg, fmt="0.0000"))
        if fuel_cost:
            ws.cell(row=i, column=5).number_format = "0.0000"
        _apply(ws, i, 6, _cell(d.get("date", ""), bg=bg))
        src = d.get("source", "")
        _apply(ws, i, 7, _cell(src, bg=bg, align="left",
                               color="C00000" if src == "估算" else "000000"))
        _apply(ws, i, 8, _cell(d.get("note", ""), bg=bg, align="left"))
        ws.row_dimensions[i].height = 18

    note_row = 3 + len(coal_data.get("data", []))
    ws.merge_cells(f"A{note_row}:H{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="折算公式：度电燃料成本 = 煤价(元/吨) × [305 × 7000 ÷ 实际热值(kcal/kg)] ÷ 1,000,000。"
                      "在线抓取失败时使用内置数据。")
    n.font = Font(name="微软雅黑", size=9, color="595959", italic=True)
    n.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 24


# ── 主函数 ────────────────────────────────────────────────
def generate_excel(spot_data: dict, coal_data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()

    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()

    build_summary(ws1, spot_data)
    build_spot(ws2, spot_data)
    build_mechanism(ws3)
    build_tou(ws4)
    build_monthly_contract(ws5)
    build_coal_prices(ws6, coal_data)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[excel] 已保存: {output_path}")
    return output_path
