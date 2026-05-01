"""
生成电力公司研究 Excel，每家公司一个文件，含4个工作表：
  Sheet1 装机汇总   - 分省煤/燃气/风/光装机及合计
  Sheet2 煤电&燃机  - 每台机组详情+灵活性改造情况
  Sheet3 新能源     - 分省风电/光伏装机
  Sheet4 发电量     - 最新发电量情况
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from companies import COMPANIES

C_HEADER_BG  = "1F4E79"
C_HEADER_FG  = "FFFFFF"
C_ODD        = "EBF3FB"
C_EVEN       = "FFFFFF"
C_WARN       = "FFE0E0"
C_GOOD       = "E2EFDA"
C_PENDING    = "FFF2CC"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(text, size=10):
    return {
        "value": text,
        "font": Font(name="微软雅黑", bold=True, size=size, color=C_HEADER_FG),
        "fill": PatternFill("solid", fgColor=C_HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin(),
    }


def _cell(value, bg=None, bold=False, color="000000", align="center", wrap=False, fmt=None):
    d = {
        "value": value,
        "font": Font(name="微软雅黑", bold=bold, size=10, color=color),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap),
        "border": _thin(),
    }
    if bg:
        d["fill"] = PatternFill("solid", fgColor=bg)
    if fmt:
        d["number_format"] = fmt
    return d


def _apply(ws, row, col, props):
    c = ws.cell(row=row, column=col)
    c.value = props.get("value")
    for attr in ("font", "fill", "alignment", "border"):
        if attr in props:
            setattr(c, attr, props[attr])
    if "number_format" in props:
        c.number_format = props["number_format"]
    return c


def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _title(ws, text, cols, size=13):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="微软雅黑", bold=True, size=size, color=C_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def _pending_or(val, fallback="—"):
    if val is None:
        return "待更新"
    return val if val != "" else fallback


# ── Sheet 1：装机汇总 ─────────────────────────────────────────
def build_capacity_summary(ws, company_name, data):
    ws.title = "装机汇总"
    ws.sheet_view.showGridLines = False
    update_time = datetime.now().strftime("%Y年%m月%d日")

    _title(ws, f"{company_name} 分省装机汇总  |  数据更新：{update_time}", 7)

    headers = ["省份", "煤电装机\n(MW)", "燃机装机\n(MW)", "风电装机\n(MW)", "光伏装机\n(MW)", "合计\n(MW)", "备注"]
    widths  = [12, 14, 14, 14, 14, 14, 20]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h))
        _w(ws, ci, w)
    ws.row_dimensions[2].height = 42

    total_coal = total_gas = total_wind = total_solar = 0
    row = 3
    for prov, pdata in data["provinces"].items():
        bg = C_ODD if row % 2 == 1 else C_EVEN
        coal_mw  = sum(u["mw"] for u in pdata.get("coal_units", []))
        gas_mw   = sum(u["mw"] for u in pdata.get("gas_units", []))
        wind_mw  = pdata.get("wind_mw") or 0
        solar_mw = pdata.get("solar_mw") or 0
        total_mw = coal_mw + gas_mw + wind_mw + solar_mw

        total_coal  += coal_mw
        total_gas   += gas_mw
        total_wind  += wind_mw
        total_solar += solar_mw

        _apply(ws, row, 1, _cell(prov, bg=bg, bold=True, align="left"))
        _apply(ws, row, 2, _cell(coal_mw  or "—", bg=bg))
        _apply(ws, row, 3, _cell(gas_mw   or "—", bg=bg))
        _apply(ws, row, 4, _cell(wind_mw  or "—", bg=bg))
        _apply(ws, row, 5, _cell(solar_mw or "—", bg=bg))
        _apply(ws, row, 6, _cell(total_mw or "—", bg=bg, bold=True))
        _apply(ws, row, 7, _cell("", bg=bg))
        ws.row_dimensions[row].height = 18
        row += 1

    # 合计行
    grand = total_coal + total_gas + total_wind + total_solar
    bg = "D9E1F2"
    _apply(ws, row, 1, _cell("合计", bg=bg, bold=True))
    for ci, val in enumerate([total_coal, total_gas, total_wind, total_solar, grand], 2):
        _apply(ws, row, ci, _cell(val or "—", bg=bg, bold=True))
    _apply(ws, row, 7, _cell("", bg=bg))
    ws.row_dimensions[row].height = 20

    # 公司整体发电量
    gen = data.get("company_generation", {})
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1,
                value=f"【发电量】{gen.get('period','—')}：总计 {gen.get('total_twh','—')} 亿kWh"
                      f"（火电 {gen.get('thermal_twh','—')}，可再生 {gen.get('renewable_twh','—')}）"
                      f"  来源：{gen.get('source','—')}")
    c.font = Font(name="微软雅黑", size=10, color="1F4E79", bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22

    # ── 装机数据完整度校验 ──────────────────────────────────────
    rep = data.get("reported_capacity", {})
    if rep:
        row += 2
        # 校验区标题行
        ws.merge_cells(f"A{row}:G{row}")
        ch = ws.cell(row=row, column=1,
                     value=f"【装机数据完整度校验】  对照：{rep.get('source','年报')}  ({rep.get('year','')}年)")
        ch.font = Font(name="微软雅黑", bold=True, size=10, color=C_HEADER_FG)
        ch.fill = PatternFill("solid", fgColor="2E4057")
        ch.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # 校验列头
        val_headers = ["类型", "本表已录入(MW)", "年报公布(MW)", "覆盖率", "差距(MW)", "状态", "备注"]
        for ci, h in enumerate(val_headers, 1):
            _apply(ws, row, ci, _hdr(h, size=9))
        ws.row_dimensions[row].height = 18
        row += 1

        checks = [
            ("煤电",   total_coal,  rep.get("coal_mw")),
            ("燃气",   total_gas,   rep.get("gas_mw")),
            ("风电",   total_wind,  rep.get("wind_mw")),
            ("光伏",   total_solar, rep.get("solar_mw")),
        ]
        compiled_total = total_coal + total_gas + total_wind + total_solar
        rep_total = (rep.get("coal_mw") or 0) + (rep.get("gas_mw") or 0) + \
                    (rep.get("wind_mw") or 0) + (rep.get("solar_mw") or 0)

        for fuel_type, compiled, reported in checks:
            bg = C_ODD if row % 2 == 1 else C_EVEN
            if reported:
                pct = compiled / reported * 100
                gap = reported - compiled
                if pct >= 80:
                    status, status_bg = "✔ 基本完整", C_GOOD
                elif pct >= 50:
                    status, status_bg = "⚠ 部分缺失", C_PENDING
                else:
                    status, status_bg = "✘ 大量缺失", C_WARN
                pct_str = f"{pct:.1f}%"
            else:
                gap, pct_str, status, status_bg = "—", "—", "年报无数据", bg

            _apply(ws, row, 1, _cell(fuel_type, bg=bg, bold=True))
            _apply(ws, row, 2, _cell(compiled,  bg=bg))
            _apply(ws, row, 3, _cell(reported if reported else "—", bg=bg))
            _apply(ws, row, 4, _cell(pct_str,   bg=bg))
            _apply(ws, row, 5, _cell(gap,        bg=bg))
            _apply(ws, row, 6, _cell(status,     bg=status_bg,
                                     color="375623" if status_bg == C_GOOD
                                     else ("7F6000" if status_bg == C_PENDING else "C00000")))
            _apply(ws, row, 7, _cell("",          bg=bg))
            ws.row_dimensions[row].height = 18
            row += 1

        # 合计校验行
        pct_tot = compiled_total / rep_total * 100 if rep_total else 0
        gap_tot  = rep_total - compiled_total
        bg = "D9E1F2"
        _apply(ws, row, 1, _cell("合计",           bg=bg, bold=True))
        _apply(ws, row, 2, _cell(compiled_total,   bg=bg, bold=True))
        _apply(ws, row, 3, _cell(rep.get("total_mw", rep_total), bg=bg, bold=True))
        _apply(ws, row, 4, _cell(f"{pct_tot:.1f}%", bg=bg, bold=True))
        _apply(ws, row, 5, _cell(round(gap_tot, 1),  bg=bg, bold=True))
        if pct_tot >= 80:
            tot_status, tot_sc = "✔ 基本完整", "375623"
        elif pct_tot >= 50:
            tot_status, tot_sc = "⚠ 部分缺失", "7F6000"
        else:
            tot_status, tot_sc = "✘ 大量缺失", "C00000"
        _apply(ws, row, 6, _cell(tot_status, bg=bg, bold=True, color=tot_sc))
        _apply(ws, row, 7, _cell("", bg=bg))
        ws.row_dimensions[row].height = 20


# ── Sheet 2：煤电&燃机明细 ─────────────────────────────────────
def build_thermal_units(ws, company_name, data):
    ws.title = "煤电&燃机明细"
    ws.sheet_view.showGridLines = False

    _title(ws, f"{company_name} 煤电及燃机机组明细（含灵活性改造）", 11)

    headers = ["省份", "电厂", "机组", "装机\n(MW)", "机型",
               "投产\n年份", "状态", "灵活性\n改造",
               "最大调峰\n能力(MW)", "供电煤耗\n(g/kWh)", "燃料\n类型", "备注"]
    widths  = [10, 18, 10, 10, 12, 8, 8, 10, 14, 14, 8, 20]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h))
        _w(ws, ci, w)
    ws.row_dimensions[2].height = 48

    row = 3
    for prov, pdata in data["provinces"].items():
        units = pdata.get("coal_units", []) + pdata.get("gas_units", [])
        if not units:
            continue
        for u in units:
            bg = C_ODD if row % 2 == 1 else C_EVEN
            fuel = "燃气" if u in pdata.get("gas_units", []) else "煤"

            flex = u.get("flexible")
            if flex is True:
                flex_val, flex_bg = "✔ 已改造", C_GOOD
            elif flex is False:
                flex_val, flex_bg = "✘ 未改造", C_WARN
            else:
                flex_val, flex_bg = "待核实", C_PENDING

            peak = u.get("peak_cap_mw")
            rate = u.get("coal_rate")
            status = u.get("status", "在运")
            status_color = "595959" if status != "在运" else "000000"

            _apply(ws, row, 1,  _cell(prov,                  bg=bg, bold=True))
            _apply(ws, row, 2,  _cell(u.get("plant",""),     bg=bg, align="left"))
            _apply(ws, row, 3,  _cell(u.get("unit",""),      bg=bg))
            _apply(ws, row, 4,  _cell(u.get("mw"),           bg=bg))
            _apply(ws, row, 5,  _cell(u.get("type",""),      bg=bg))
            _apply(ws, row, 6,  _cell(u.get("cod"),          bg=bg))
            _apply(ws, row, 7,  _cell(status,                bg=bg, color=status_color))
            _apply(ws, row, 8,  _cell(flex_val,              bg=flex_bg,
                                      color="375623" if flex else ("C00000" if flex is False else "7F6000")))
            _apply(ws, row, 9,  _cell(peak if peak else "待更新",  bg=bg,
                                      color="000000" if peak else "999999"))
            _apply(ws, row, 10, _cell(rate if rate else "待更新",  bg=bg,
                                      color="000000" if rate else "999999"))
            _apply(ws, row, 11, _cell(fuel,                  bg=bg))
            _apply(ws, row, 12, _cell(u.get("note",""),      bg=bg, align="left", wrap=True))
            ws.row_dimensions[row].height = 18
            row += 1

    if row == 3:
        ws.merge_cells(f"A3:L3")
        ws.cell(row=3, column=1, value="暂无煤电/燃机机组数据，待 Remote Agent 更新").font = \
            Font(name="微软雅黑", size=10, color="999999", italic=True)

    note_row = row + 1
    ws.merge_cells(f"A{note_row}:L{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="灵活性改造：✔绿色=已完成，✘红色=未改造，黄色=待核实。"
                      "调峰能力=最低稳燃负荷（MW）。供电煤耗含厂用电。标注\"待核实\"的数据需 Remote Agent 从年报/公告核实。")
    n.font = Font(name="微软雅黑", size=9, color="595959", italic=True)
    n.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 28


# ── Sheet 3：新能源装机 ───────────────────────────────────────
def build_renewables(ws, company_name, data):
    ws.title = "新能源装机"
    ws.sheet_view.showGridLines = False

    _title(ws, f"{company_name} 分省新能源装机", 5)

    headers = ["省份", "风电装机(MW)", "光伏装机(MW)", "合计(MW)", "备注"]
    widths  = [12, 16, 16, 14, 24]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h))
        _w(ws, ci, w)
    ws.row_dimensions[2].height = 36

    total_wind = total_solar = 0
    row = 3
    for prov, pdata in data["provinces"].items():
        wind  = pdata.get("wind_mw")  or 0
        solar = pdata.get("solar_mw") or 0
        if wind == 0 and solar == 0:
            continue
        bg = C_ODD if row % 2 == 1 else C_EVEN
        total_wind  += wind
        total_solar += solar

        _apply(ws, row, 1, _cell(prov,        bg=bg, bold=True, align="left"))
        _apply(ws, row, 2, _cell(wind or "—", bg=bg))
        _apply(ws, row, 3, _cell(solar or "—", bg=bg))
        _apply(ws, row, 4, _cell((wind + solar) or "—", bg=bg, bold=True))
        _apply(ws, row, 5, _cell("",           bg=bg))
        ws.row_dimensions[row].height = 18
        row += 1

    bg = "D9E1F2"
    _apply(ws, row, 1, _cell("合计", bg=bg, bold=True))
    _apply(ws, row, 2, _cell(total_wind,  bg=bg, bold=True))
    _apply(ws, row, 3, _cell(total_solar, bg=bg, bold=True))
    _apply(ws, row, 4, _cell(total_wind + total_solar, bg=bg, bold=True))
    _apply(ws, row, 5, _cell("", bg=bg))
    ws.row_dimensions[row].height = 20


# ── Sheet 4：发电量 ──────────────────────────────────────────
def build_generation(ws, company_name, data):
    ws.title = "发电量"
    ws.sheet_view.showGridLines = False

    _title(ws, f"{company_name} 发电量情况  |  更新日期：{datetime.now().strftime('%Y-%m-%d')}", 7)

    headers = ["省份", "统计期", "总发电量\n(亿kWh)", "火电\n(亿kWh)", "新能源\n(亿kWh)", "同比(%)", "数据来源"]
    widths  = [12, 12, 16, 14, 14, 12, 24]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws, 2, ci, _hdr(h))
        _w(ws, ci, w)
    ws.row_dimensions[2].height = 48

    row = 3
    for prov, pdata in data["provinces"].items():
        bg = C_ODD if row % 2 == 1 else C_EVEN
        gen = pdata.get("generation", {})
        period = gen.get("period") or "—"
        total  = gen.get("total_twh")
        therm  = gen.get("thermal_twh")
        renew  = gen.get("renewable_twh")
        yoy    = gen.get("yoy")
        src    = gen.get("source") or "待更新"

        _apply(ws, row, 1, _cell(prov,                              bg=bg, bold=True, align="left"))
        _apply(ws, row, 2, _cell(period,                            bg=bg))
        _apply(ws, row, 3, _cell(total  if total  is not None else "待更新", bg=bg if total  is not None else C_PENDING,
                                 color="000000" if total is not None else "7F6000"))
        _apply(ws, row, 4, _cell(therm  if therm  is not None else "待更新", bg=bg if therm  is not None else C_PENDING,
                                 color="000000" if therm is not None else "7F6000"))
        _apply(ws, row, 5, _cell(renew  if renew  is not None else "待更新", bg=bg if renew  is not None else C_PENDING,
                                 color="000000" if renew is not None else "7F6000"))
        _apply(ws, row, 6, _cell(f"{yoy:+.1f}%" if yoy is not None else "—", bg=bg,
                                 color="375623" if (yoy and yoy > 0) else ("C00000" if (yoy and yoy < 0) else "999999")))
        _apply(ws, row, 7, _cell(src, bg=bg, align="left",
                                 color="999999" if src == "待更新" else "000000"))
        ws.row_dimensions[row].height = 18
        row += 1

    # 公司整体行
    row += 1
    gen = data.get("company_generation", {})
    bg = "D9E1F2"
    _apply(ws, row, 1, _cell("全公司合计", bg=bg, bold=True, align="left"))
    _apply(ws, row, 2, _cell(gen.get("period","—"), bg=bg))
    _apply(ws, row, 3, _cell(gen.get("total_twh","待更新"),   bg=bg, bold=True))
    _apply(ws, row, 4, _cell(gen.get("thermal_twh","待更新"), bg=bg, bold=True))
    _apply(ws, row, 5, _cell(gen.get("renewable_twh","待更新"), bg=bg, bold=True))
    yoy = gen.get("yoy_total")
    _apply(ws, row, 6, _cell(f"{yoy:+.1f}%" if yoy is not None else "—", bg=bg))
    _apply(ws, row, 7, _cell(gen.get("source","待更新"), bg=bg, align="left"))
    ws.row_dimensions[row].height = 20


# ── 主函数 ───────────────────────────────────────────────────
def generate_company_excel(company_name: str, output_dir: str) -> str:
    data = COMPANIES.get(company_name)
    if not data:
        raise ValueError(f"未知公司：{company_name}")

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(output_dir, f"{company_name}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws2, ws3, ws4 = wb.create_sheet(), wb.create_sheet(), wb.create_sheet()

    build_capacity_summary(ws1, company_name, data)
    build_thermal_units(ws2, company_name, data)
    build_renewables(ws3, company_name, data)
    build_generation(ws4, company_name, data)

    wb.save(path)
    print(f"[company] 已保存：{path}")
    return path
